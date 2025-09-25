import argparse
import io
import os
from pathlib import Path
from typing import List, Tuple, Optional

import numpy as np
from PIL import Image

import torch
import torch.nn as nn
from torchvision import transforms
from torchvision.models import efficientnet_b0, EfficientNet_B0_Weights

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
import zipfile

# ----------------------------
# 画像探索
# ----------------------------
IMG_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}

def list_images(root: Path) -> List[Path]:
    files = []
    for p in root.rglob("*"):
        if p.is_file() and p.suffix.lower() in IMG_EXTS:
            files.append(p)
    files.sort()
    return files

# ----------------------------
# ラベル読み込み
# ----------------------------
def load_labels(labels_path: Optional[Path], num_classes: int) -> List[str]:
    if labels_path and labels_path.exists():
        with open(labels_path, encoding="utf-8") as f:
            labels = [line.strip() for line in f if line.strip()]
        return labels
    return [str(i) for i in range(num_classes)]

# ----------------------------
# モデル読み込み（EfficientNet-B0, 256x256）
# ----------------------------
def build_model(num_classes: int):
    """
    ImageNet事前学習のEfficientNet-B0をベースに最終層を付け替え。
    """
    # weights=None でもOKだが、finetune前提なら初期形はどちらでも可
    model = efficientnet_b0(weights=None)
    in_features = model.classifier[1].in_features
    model.classifier[1] = nn.Linear(in_features, num_classes)
    return model

def flexible_load_state_dict(model: nn.Module, state_dict: dict):
    """state_dictのキー違いに柔軟対応"""
    try:
        model.load_state_dict(state_dict, strict=True)
        return
    except Exception:
        pass
    # 'model'キーに入っているケース
    if "model" in state_dict and isinstance(state_dict["model"], dict):
        try:
            model.load_state_dict(state_dict["model"], strict=False)
            return
        except Exception:
            pass
    # 'state_dict'に入っているケース
    if "state_dict" in state_dict and isinstance(state_dict["state_dict"], dict):
        try:
            model.load_state_dict(state_dict["state_dict"], strict=False)
            return
        except Exception:
            pass
    # 余計な接頭辞を外す
    new_sd = {}
    for k, v in state_dict.items():
        nk = k
        if nk.startswith("module."):
            nk = nk[len("module."):]
        new_sd[nk] = v
    model.load_state_dict(new_sd, strict=False)

def load_model(model_path: Path, num_classes: int, device: torch.device) -> nn.Module:
    model = build_model(num_classes)
    # .za（zip系）にも対応：中の .pth / .pt を探してロード
    if model_path.suffix.lower() in [".za", ".zip"]:
        with zipfile.ZipFile(model_path, 'r') as zf:
            # 最初に見つかった .pth/.pt を使う
            cand = None
            for n in zf.namelist():
                if n.lower().endswith((".pth", ".pt", ".pth.tar")):
                    cand = n
                    break
            if cand is None:
                raise ValueError("zip内に .pth/.pt が見つかりません。")
            with zf.open(cand, 'r') as f:
                buffer = f.read()
            state = torch.load(io.BytesIO(buffer), map_location="cpu")
    else:
        state = torch.load(model_path, map_location="cpu")
    flexible_load_state_dict(model, state)
    model.eval()
    model.to(device)
    return model

# ----------------------------
# 前処理＆推論
# ----------------------------
def make_transform(input_size: Tuple[int, int]):
    # 256x256 指定に合わせてリサイズ固定（CenterCrop不要）
    # ImageNetの平均・分散で正規化
    H, W = input_size
    return transforms.Compose([
        transforms.Resize((H, W)),
        transforms.ToTensor(),
        transforms.Normalize(mean=EfficientNet_B0_Weights.IMAGENET1K_V1.transforms.mean,
                             std=EfficientNet_B0_Weights.IMAGENET1K_V1.transforms.std),
    ])

def predict_probs(model: nn.Module,
                  img_paths: List[Path],
                  tfm,
                  batch_size: int,
                  device: torch.device) -> np.ndarray:
    probs_all = []
    batch_imgs = []
    with torch.no_grad():
        for p in img_paths:
            img = Image.open(p).convert("RGB")
            x = tfm(img)  # [3,H,W]
            batch_imgs.append(x)
            if len(batch_imgs) == batch_size:
                X = torch.stack(batch_imgs, dim=0).to(device)  # [B,3,H,W]
                logits = model(X)                              # [B,C]
                probs = torch.softmax(logits, dim=1).cpu().numpy()
                probs_all.append(probs)
                batch_imgs = []
        if batch_imgs:
            X = torch.stack(batch_imgs, dim=0).to(device)
            logits = model(X)
            probs = torch.softmax(logits, dim=1).cpu().numpy()
            probs_all.append(probs)
    return np.vstack(probs_all)

# ----------------------------
# Excel 出力（画像サムネ埋め込み）
# ----------------------------
def image_to_xlimage(pil_img: Image.Image, thumb_max: int = 128) -> XLImage:
    w, h = pil_img.size
    scale = min(thumb_max / max(w, h), 1.0)
    if scale < 1.0:
        pil_img = pil_img.resize((int(w*scale), int(h*scale)))
    bio = io.BytesIO()
    pil_img.save(bio, format="PNG")
    bio.seek(0)
    return XLImage(bio)

def write_excel(rows: List[dict], out_xlsx: Path, thumb_max: int = 128):
    wb = Workbook()
    ws = wb.active
    ws.title = "inference"
    header = [
        "フォルダ", "画像名", "画像表示",
        "推論top1ラベル", "top1精度",
        "推論top2ラベル", "top2精度",
        "推論top3ラベル", "top3精度",
        "推論top4ラベル", "top4精度",
        "自動/目視"
    ]
    ws.append(header)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for c in range(1, len(header)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions["C"].width = 24

    row_idx = 2
    for r in rows:
        ws.append([
            r["folder"],
            r["filename"],
            "",
            r["top1_label"], r["top1_score"],
            r["top2_label"], r["top2_score"],
            r["top3_label"], r["top3_score"],
            r["top4_label"], r["top4_score"],
            r["auto_or_manual"],
        ])
        # サムネ貼り付け
        xlimg = image_to_xlimage(r["pil_image"], thumb_max=thumb_max)
        ws.add_image(xlimg, f"C{row_idx}")
        ws.row_dimensions[row_idx].height = max(thumb_max * 0.75, 20)
        for col in ["D","F","H","J","L"]:
            ws[f"{col}{row_idx}"].alignment = Alignment(horizontal="center")
        row_idx += 1
    wb.save(out_xlsx)

# ----------------------------
# 自動/目視 判定
# ----------------------------
def decide_auto_or_manual(top1: float, top2: float, th_score: float, th_margin: float) -> str:
    return "自動" if (top1 >= th_score and (top1 - top2) >= th_margin) else "目視"

# ----------------------------
# メイン
# ----------------------------
def main():
    ap = argparse.ArgumentParser(description="EfficientNetB0(286cls) 一括推論→Excel(サムネ/Top1-4/自動-目視)")
    ap.add_argument("--root", required=True, help="annotation_images のルート")
    ap.add_argument("--out", default="inference_results.xlsx", help="出力Excelパス")
    ap.add_argument("--labels", default=None, help="ラベルファイル（1行1ラベル）")
    ap.add_argument("--num_classes", type=int, default=286)
    ap.add_argument("--batch_size", type=int, default=128)
    ap.add_argument("--input_size", type=int, nargs=2, default=[256, 256], help="H W")
    ap.add_argument("--th_score", type=float, default=0.90)
    ap.add_argument("--th_margin", type=float, default=0.20)
    ap.add_argument("--model", default="best_acc_mode.pth.za", help="モデルファイル（.pth/.pt も可、.za/.zip対応）")
    ap.add_argument("--thumb", type=int, default=128, help="Excelサムネ最大辺(px)")
    ap.add_argument("--device", default=None, help="cuda:0 等。未指定なら自動選択")
    args = ap.parse_args()

    root = Path(args.root)
    out_xlsx = Path(args.out)

    device_str = args.device if args.device else ("cuda:0" if torch.cuda.is_available() else "cpu")
    device = torch.device(device_str)
    print(f"[Device] {device}")

    labels = load_labels(Path(args.labels) if args.labels else None, args.num_classes)
    model = load_model(Path(args.model), args.num_classes, device)
    tfm = make_transform(tuple(args.input_size))

    # 画像一覧
    img_paths = list_images(root)
    if not img_paths:
        print("画像が見つかりませんでした。")
        return
    print(f"[Info] 画像枚数: {len(img_paths)}")

    # 推論
    probs = predict_probs(model, img_paths, tfm, args.batch_size, device)
    C = probs.shape[1]
    assert C == len(labels), f"ラベル数({len(labels)})とモデル出力({C})が不一致です。"

    # Top-4
    topk = 4
    idx_sorted = np.argsort(-probs, axis=1)[:, :topk]
    scores_sorted = np.take_along_axis(probs, idx_sorted, axis=1)

    # Excel用の行データ生成（PILはExcel埋め込みに使う）
    rows = []
    for i, p in enumerate(img_paths):
        # 表示用にPILも保持（別読み）
        try:
            pil_img = Image.open(p).convert("RGB")
        except Exception as e:
            print(f"(サムネ読み込み失敗) {p}: {e}")
            pil_img = Image.new("RGB", (64, 64), (200, 200, 200))

        folder_rel = str(p.parent.relative_to(root))
        filename = p.name
        idxs = idx_sorted[i].tolist()
        scs = scores_sorted[i].tolist()

        # guard
        while len(idxs) < 4:
            idxs.append(idxs[-1])
            scs.append(0.0)

        top1, top2 = scs[0], scs[1]
        auto_or_manual = decide_auto_or_manual(top1, top2, args.th_score, args.th_margin)

        rows.append({
            "folder": folder_rel,
            "filename": filename,
            "pil_image": pil_img,
            "top1_label": labels[idxs[0]],
            "top1_score": float(f"{scs[0]:.4f}"),
            "top2_label": labels[idxs[1]],
            "top2_score": float(f"{scs[1]:.4f}"),
            "top3_label": labels[idxs[2]],
            "top3_score": float(f"{scs[2]:.4f}"),
            "top4_label": labels[idxs[3]],
            "top4_score": float(f"{scs[3]:.4f}"),
            "auto_or_manual": auto_or_manual
        })

    write_excel(rows, out_xlsx, thumb_max=args.thumb)
    print(f"[Done] Excel 保存: {out_xlsx}")

if __name__ == "__main__":
    main()

python infer_to_excel_efficientnet.py \
  --root "annotation_images" \
  --out "inference_results.xlsx" \
  --labels "classes_286.txt" \
  --num_classes 286 \
  --batch_size 128 \
  --input_size 256 256 \
  --model "best_acc_mode.pth.za" \
  --th_score 0.95 \
  --th_margin 0.25
