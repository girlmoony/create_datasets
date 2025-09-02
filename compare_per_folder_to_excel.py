import argparse
import os
import sys
import shlex
from pathlib import Path
from typing import Dict, Tuple, Optional, List
import hashlib
from collections import defaultdict

import paramiko
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ============== SSH 接続 ==============

def ssh_connect(host: str, user: str, port: int, key_path: Optional[str], password: Optional[str]) -> paramiko.SSHClient:
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    pkey = None
    if key_path:
        key_path = os.path.expanduser(key_path)
        try:
            try:
                pkey = paramiko.Ed25519Key.from_private_key_file(key_path)
            except Exception:
                pkey = paramiko.RSAKey.from_private_key_file(key_path)
        except Exception as e:
            print(f"[ERROR] Private key load failed: {e}", file=sys.stderr)
            sys.exit(2)
    client.connect(hostname=host, username=user, port=port, pkey=pkey, password=password, look_for_keys=(pkey is None))
    return client


# ============== スキャン ==============

def scan_local_files(root: Path) -> Dict[str, Tuple[int, float]]:
    """
    ローカル A: 相対パス -> (サイズ, mtime)
    """
    mapping: Dict[str, Tuple[int, float]] = {}
    root = root.resolve()
    for dirpath, _, filenames in os.walk(root):
        base = Path(dirpath)
        for fn in filenames:
            p = base / fn
            try:
                st = p.stat()
                rel = str(p.relative_to(root)).replace("\\", "/")
                mapping[rel] = (int(st.st_size), float(st.st_mtime))
            except Exception:
                continue
    return mapping


def scan_remote_files_metadata(ssh: paramiko.SSHClient, remote_root: str) -> Dict[str, Tuple[int, Optional[float]]]:
    """
    リモート B: 相対パス -> (サイズ, mtime or None)
    GNU find の -printf を使って「path<TAB>size<TAB>mtime<NUL>」を優先, 無ければフォールバック。
    """
    mapping: Dict[str, Tuple[int, Optional[float]]] = {}
    remote_root_q = shlex.quote(remote_root)
    cmd = (
        f'cd {remote_root_q} && '
        f'(find . -type f -printf "%P\\t%s\\t%T@\\0" 2>/dev/null || '
        f' find . -type f -printf "%P\\t%s\\0" 2>/dev/null || '
        f' find . -type f -print0)'
    )
    stdin, stdout, stderr = ssh.exec_command(cmd)
    stdin.close()

    buf = b""
    while True:
        chunk = stdout.channel.recv(32768)
        if not chunk:
            break
        buf += chunk
        parts = buf.split(b"\x00")
        buf = parts[-1]
        for raw in parts[:-1]:
            if not raw:
                continue
            s = raw.decode("utf-8", errors="replace")
            segs = s.split("\t")
            if len(segs) == 3:
                rel, size_s, mtime_s = segs
                size = int(size_s) if size_s.isdigit() else -1
                try:
                    mtime = float(mtime_s)
                except Exception:
                    mtime = None
            elif len(segs) == 2:
                rel, size_s = segs
                size = int(size_s) if size_s.isdigit() else -1
                mtime = None
            else:
                rel = segs[0]
                size = -1
                mtime = None

            rel = rel.strip().replace("\\", "/")
            if rel.startswith("./"):
                rel = rel[2:]
            if rel:
                mapping[rel] = (size, mtime)

    err = stderr.read().decode("utf-8", errors="replace").strip()
    if err:
        print(f"[WARN] remote scan stderr: {err}", file=sys.stderr)
    return mapping


# ============== ハッシュ関連（必要時） ==============

def detect_remote_hash_cmd(ssh: paramiko.SSHClient, preferred: str) -> Optional[str]:
    """sha1sum/md5sum/sha256sum のいずれかを優先順位で探す"""
    cmd_map = {"sha1": "sha1sum", "md5": "md5sum", "sha256": "sha256sum"}
    candidates = [cmd_map.get(preferred, "sha1sum"), "sha1sum", "md5sum", "sha256sum"]
    for c in candidates:
        _in, out, _err = ssh.exec_command(f"command -v {c} >/dev/null 2>&1; echo $?")
        if out.read().decode().strip() == "0":
            return c
    return None


def local_file_hash(path: Path, algo: str) -> str:
    h = hashlib.new(algo)
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def scan_remote_hashes_subset(ssh: paramiko.SSHClient, remote_root: str, algo_cmd: str, rel_paths: List[str]) -> Dict[str, str]:
    """
    in_both の相対パス群だけを対象に、ハッシュを取得。
    パスは改行区切りで安全に渡す（NUL対応の sum が無いので xargs -d '\\n' を使う）。
    """
    if not rel_paths:
        return {}
    remote_root_q = shlex.quote(remote_root)
    joined = "\n".join(f"./{p}" for p in rel_paths)  # sha1sum 等は "./path" 前提でもOK
    # 注意: BSD xargs では -d が無いことがある。GNU 環境推奨。互換性を考慮しつつ標準的ケースを想定。
    cmd = f'cd {remote_root_q} && printf "%s\\n" {shlex.quote(joined)} | xargs -d "\\n" {algo_cmd}'
    stdin, stdout, stderr = ssh.exec_command(cmd)
    stdin.close()

    mapping: Dict[str, str] = {}
    for line_b in stdout:
        line = line_b.decode("utf-8", errors="replace").rstrip("\n")
        if not line:
            continue
        try:
            h, p = line.split(None, 1)
        except ValueError:
            continue
        p = p.strip()
        if p.startswith("./"):
            p = p[2:]
        p = p.replace("\\", "/")
        mapping[p] = h

    err = stderr.read().decode("utf-8", errors="replace").strip()
    if err:
        print(f"[WARN] remote hash subset stderr: {err}", file=sys.stderr)
    return mapping


# ============== フォルダキー/集計 ==============

def folder_key_for_level(relpath: str, level: int) -> str:
    """ファイル相対パスから、先頭 level 階層のフォルダキーを作成"""
    relpath = relpath.strip().replace("\\", "/")
    parts = [p for p in relpath.split("/") if p]
    if len(parts) <= 1 or level <= 0:
        return "(root)"
    return "/".join(parts[:-1][:level]) or "(root)"


# ============== Excel 出力 ==============

def write_excel(
    folders_df: pd.DataFrame,
    only_a_df: pd.DataFrame,
    only_b_df: pd.DataFrame,
    diff_df: pd.DataFrame,
    out_path: Path
):
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        folders_df.to_excel(w, index=False, sheet_name="folders")
        only_a_df.to_excel(w, index=False, sheet_name="only_in_A")
        only_b_df.to_excel(w, index=False, sheet_name="only_in_B")
        diff_df.to_excel(w, index=False, sheet_name="diff_in_both")

    # 体裁整え
    wb = load_workbook(out_path)

    def style_ws(name: str, widths: Dict[str, int]):
        ws = wb[name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    style_ws("folders",   {"A": 40, "B": 14, "C": 14, "D": 14, "E": 14, "F": 16, "G": 16, "H": 22})
    style_ws("only_in_A", {"A": 40, "B": 60, "C": 14, "D": 20})
    style_ws("only_in_B", {"A": 40, "B": 60, "C": 14, "D": 20})
    # diff は method によって列が増えることがあるが大体このくらい
    if "diff_in_both" in wb.sheetnames:
        ws = wb["diff_in_both"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        # 適当に広めに
        for col_letter in ["A","B","C","D","E","F"]:
            ws.column_dimensions[col_letter].width = 40

    wb.save(out_path)


# ============== メイン ==============

def main():
    ap = argparse.ArgumentParser(description="Per-folder diff between local A and remote B (SSH), export to Excel with lists only when different.")
    ap.add_argument("--local_a", required=True, help="Path to local A")
    ap.add_argument("--remote_host", required=True, help="SSH host")
    ap.add_argument("--remote_user", required=True, help="SSH username")
    ap.add_argument("--remote_b", required=True, help="Path to remote B")
    ap.add_argument("--port", type=int, default=22, help="SSH port (default: 22)")
    ap.add_argument("--key", default=None, help="Private key path (optional)")
    ap.add_argument("--password", default=None, help="Password (optional)")
    ap.add_argument("--level", type=int, default=1, help="Aggregation folder level (1=direct subfolders)")
    ap.add_argument("--method", choices=["metadata", "hash"], default="metadata", help="How to detect 'different content'")
    ap.add_argument("--hash_algo", choices=["sha1", "md5", "sha256"], default="sha1", help="Hash algorithm when --method hash")
    ap.add_argument("--out", default="compare_report.xlsx", help="Output Excel path")
    args = ap.parse_args()

    local_root = Path(args.local_a)
    if not local_root.exists() or not local_root.is_dir():
        print(f"[ERROR] Local A not found or not a directory: {local_root}", file=sys.stderr)
        sys.exit(1)

    # スキャン
    print("[INFO] Scanning local A ...")
    local_meta = scan_local_files(local_root)

    print("[INFO] Scanning remote B ...")
    ssh = ssh_connect(args.remote_host, args.remote_user, args.port, args.key, args.password)
    try:
        remote_meta = scan_remote_files_metadata(ssh, args.remote_b)

        # in_both のみハッシュ（必要時）
        remote_hash_cmd = None
        remote_hashes = {}
        if args.method == "hash":
            remote_hash_cmd = detect_remote_hash_cmd(ssh, args.hash_algo)
            if not remote_hash_cmd:
                print("[WARN] No sha1sum/md5sum/sha256sum on remote. Falling back to metadata method.", file=sys.stderr)
                args.method = "metadata"
    finally:
        ssh.close()

    # 差分判定
    local_paths = set(local_meta.keys())
    remote_paths = set(remote_meta.keys())

    only_in_a_paths = sorted(list(local_paths - remote_paths))
    only_in_b_paths = sorted(list(remote_paths - local_paths))
    in_both_paths = sorted(list(local_paths & remote_paths))

    # ハッシュ方式の場合は in_both のみハッシュ比較
    diff_in_both_paths: List[str] = []
    equal_in_both_count = 0

    if args.method == "metadata":
        for rel in in_both_paths:
            lsize = local_meta[rel][0]
            rsize = remote_meta[rel][0]
            if lsize >= 0 and rsize >= 0:
                if lsize != rsize:
                    diff_in_both_paths.append(rel)
                else:
                    equal_in_both_count += 1
            else:
                # サイズ不明は判定不能 → 「等しい」に加えない
                pass
    else:
        # 可能なら remote で in_both のみハッシュ計算
        if remote_hash_cmd:
            ssh = ssh_connect(args.remote_host, args.remote_user, args.port, args.key, args.password)
            try:
                remote_hashes = scan_remote_hashes_subset(ssh, args.remote_b, remote_hash_cmd, in_both_paths)
            finally:
                ssh.close()
        for rel in in_both_paths:
            r_hash = remote_hashes.get(rel)
            if not r_hash:
                # 取れなかったらスキップ（厳密に数えたい場合は全量ハッシュへ変更可）
                continue
            try:
                l_hash = local_file_hash(local_root / rel, args.hash_algo)
            except Exception:
                diff_in_both_paths.append(rel)
                continue
            if l_hash != r_hash:
                diff_in_both_paths.append(rel)
            else:
                equal_in_both_count += 1

    # フォルダキーごとの集計
    level = args.level
    def fk(p: str) -> str:
        return folder_key_for_level(p, level)

    # フォルダ別リスト（同一フォルダは一覧に出さない）
    only_a_grouped = defaultdict(list)
    only_b_grouped = defaultdict(list)
    diff_grouped   = defaultdict(list)

    for p in only_in_a_paths:
        only_a_grouped[fk(p)].append(p)
    for p in only_in_b_paths:
        only_b_grouped[fk(p)].append(p)
    for p in diff_in_both_paths:
        diff_grouped[fk(p)].append(p)

    folder_keys = sorted(set(list(only_a_grouped.keys()) + list(only_b_grouped.keys()) + list(diff_grouped.keys()) +
                             [fk(p) for p in local_paths] + [fk(p) for p in remote_paths]))

    # サマリ rows
    folder_rows = []
    for k in folder_keys:
        local_count = sum(1 for p in local_paths if fk(p) == k)
        remote_count = sum(1 for p in remote_paths if fk(p) == k)
        only_a_count = len(only_a_grouped.get(k, []))
        only_b_count = len(only_b_grouped.get(k, []))
        diff_count   = len(diff_grouped.get(k, []))
        status = "match"
        if only_a_count == 0 and only_b_count == 0 and diff_count == 0 and local_count == remote_count:
            status = "match"
        else:
            status = "mixed"
            if only_a_count > 0 and only_b_count == 0 and diff_count == 0:
                status = "only_local"
            elif only_b_count > 0 and only_a_count == 0 and diff_count == 0:
                status = "only_remote"
            elif diff_count > 0 and only_a_count == 0 and only_b_count == 0 and local_count == remote_count:
                status = "different_content"
            else:
                status = "mixed"

        folder_rows.append({
            "folder_key": k,
            "local_total_files": local_count,
            "remote_total_files": remote_count,
            "only_in_A_count": only_a_count,
            "only_in_B_count": only_b_count,
            "diff_in_both_count": diff_count,
            "diff_local_minus_remote": local_count - remote_count,
            "status": status,  # match のときは一覧に出さない
        })

    folders_df = pd.DataFrame(folder_rows).sort_values("folder_key").reset_index(drop=True)

    # 一覧シート用 DF（差分があるフォルダのみ）
    only_a_rows, only_b_rows, diff_rows = [], [], []

    for k, paths in only_a_grouped.items():
        for p in paths:
            size, mtime = local_meta[p]
            only_a_rows.append({"folder_key": k, "path": p, "size": size, "mtime": mtime})

    for k, paths in only_b_grouped.items():
        for p in paths:
            size, mtime = remote_meta[p]
            only_b_rows.append({"folder_key": k, "path": p, "size": size, "mtime": mtime})

    if args.method == "metadata":
        for k, paths in diff_grouped.items():
            for p in paths:
                diff_rows.append({
                    "folder_key": k,
                    "path": p,
                    "local_size": local_meta[p][0],
                    "remote_size": remote_meta[p][0],
                })
    else:
        # ハッシュ列も載せたければここで計算して追加可能（行数増を避けたい場合はサイズのみでも可）
        for k, paths in diff_grouped.items():
            for p in paths:
                row = {
                    "folder_key": k,
                    "path": p,
                    "local_size": local_meta[p][0],
                    "remote_size": remote_meta[p][0],
                }
                diff_rows.append(row)

    only_a_df = pd.DataFrame(only_a_rows)
    only_b_df = pd.DataFrame(only_b_rows)
    diff_df   = pd.DataFrame(diff_rows)

    out_path = Path(args.out)
    print(f"[INFO] Writing Excel to {out_path} ...")
    write_excel(folders_df, only_a_df, only_b_df, diff_df, out_path)
    print(f"[DONE] Report saved: {out_path.resolve()}")


if __name__ == "__main__":
    main()


"""
python compare_per_folder_to_excel.py \
  --local_a "/path/to/A" \
  --remote_host "example.com" \
  --remote_user "youruser" \
  --remote_b "/path/to/B" \
  --port 22 \
  --key "~/.ssh/id_ed25519" \
  --level 1 \
  --method metadata \
  --out "compare_report.xlsx"
"""
