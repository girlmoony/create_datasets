#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import argparse
import re
import shutil
from collections import defaultdict, Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import load_workbook

VALID_SUBSETS = {"train","val","test"}

@dataclass(frozen=True)
class SrcFile:
    path: Path
    store: Optional[str]

def sanitize_label(label: str) -> str:
    import re
    s = str(label or "").strip()
    s = re.sub(r'[\\/:\*\?"<>\|\r\n\t]', '_', s)
    s = re.sub(r'\s+', ' ', s)
    return s if s else "unnamed"

def resolve_src_folder(base: str, folder: str) -> Path:
    p = Path(folder)
    if p.is_absolute():
        return p
    return (Path(base) / folder).resolve()

def list_pngs(folder: Path, recursive: bool) -> List[Path]:
    if not folder.exists() or not folder.is_dir():
        return []
    if recursive:
        return list(folder.rglob("*.png"))
    else:
        return [p for p in folder.iterdir() if p.suffix.lower()==".png" and p.is_file()]

def extract_store(p: Path, pat: re.Pattern) -> Optional[str]:
    m = pat.search(p.name)
    return m.group("store") if m else None

def pick_files_balanced_then_fill(files: List[SrcFile], need: int,
                                  per_store_cap: int, existing_store_counts: Counter) -> List[SrcFile]:
    chosen: List[SrcFile] = []
    by_store: Dict[Optional[str], List[SrcFile]] = defaultdict(list)
    for f in files:
        by_store[f.store].append(f)
    for s in by_store:
        by_store[s].sort(key=lambda x: x.path.name)

    store_order = sorted(by_store.keys(), key=lambda s: (existing_store_counts.get(s, 0), str(s)))
    indices = {s: 0 for s in by_store}
    while len(chosen) < need and store_order:
        progressed = False
        for s in store_order:
            have = existing_store_counts.get(s, 0)
            if have >= per_store_cap:
                continue
            idx = indices[s]
            if idx >= len(by_store[s]):
                continue
            chosen.append(by_store[s][idx])
            indices[s] += 1
            existing_store_counts[s] = have + 1
            progressed = True
            if len(chosen) >= need:
                break
        if not progressed:
            break

    if len(chosen) >= need:
        return chosen[:need]

    chosen_set = {c.path for c in chosen}
    filler: List[SrcFile] = []
    for s in sorted(by_store.keys(), key=lambda x: str(x)):
        for f in by_store[s]:
            if f.path in chosen_set:
                continue
            filler.append(f)
            if len(chosen) + len(filler) >= need:
                break
        if len(chosen) + len(filler) >= need:
            break

    return (chosen + filler)[:need]

def main():
    ap = argparse.ArgumentParser(description="Copy N PNGs per row to --dest/<subset>/<label> without overwriting; prefer up to 6 per store, then fill.")
    ap.add_argument("--excel", required=True, help="Excel workbook path (will be updated)")
    ap.add_argument("--sheet", required=True, help="Sheet name to read/write")
    ap.add_argument("--dest", required=True, help="Destination base directory (contains train/val/test)")
    ap.add_argument("--dry-run", action="store_true", help="Plan only; write logs but do not copy files")
    ap.add_argument("--filename_regex", default=r'(?P<store>\d{3,4})', help="Regex that extracts 'store' from PNG filename")
    ap.add_argument("--recursive", action="store_true", default=True, help="Search recursively for PNGs in source folders")
    ap.add_argument("--primary_store_cap", type=int, default=6, help="Phase1 per-store cap (default 6)")
    args = ap.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")
    dest_root = Path(args.dest).resolve()
    dest_root.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(excel_path, sheet_name=args.sheet, header=0)
    cols_lower = {c.lower(): c for c in df.columns}
    count_key = None
    for k in cols_lower:
        if k in ["枚数", "count", "枚數"]:
            count_key = cols_lower[k]
            break
    required_names = ["type","product","label","base","folder"]
    missing = [n for n in required_names if n not in cols_lower]
    subset_colname = cols_lower.get("subset") or cols_lower.get("subsest")
    if subset_colname is None:
        missing.append("subset/subsest")
    if count_key is None:
        missing.append("枚数")
    if missing:
        raise ValueError(f"必要カラムが不足: {missing}  （存在: {list(df.columns)}）")

    type_col = cols_lower["type"]
    label_col = cols_lower["label"]
    base_col = cols_lower["base"]
    folder_col = cols_lower["folder"]
    subset_col = subset_colname

    df_ok = df[df[type_col].astype(str).str.upper() == "OK"].copy()
    if df_ok.empty:
        print("No rows with type == OK")
    df_ok[subset_col] = df_ok[subset_col].astype(str).str.strip().str.lower()

    def to_int(x):
        try:
            return int(float(str(x).strip()))
        except Exception:
            return 0
    df_ok["__need__"] = df_ok[count_key].apply(to_int)

    wb = load_workbook(excel_path)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {args.sheet}")
    ws = wb[args.sheet]
    header_row = 1
    log_col_idx = 7  # Column G

    store_pat = re.compile(args.filename_regex)

    dest_existing_names: Dict[Tuple[str,str], Set[str]] = defaultdict(set)
    dest_reserved_names: Dict[Tuple[str,str], Set[str]] = defaultdict(set)
    dest_store_counts: Dict[Tuple[str,str], Counter] = defaultdict(Counter)

    for idx, row in df_ok.iterrows():
        subset = str(row[subset_col]).strip().lower()
        if subset not in VALID_SUBSETS:
            ws.cell(row=header_row + 1 + idx, column=log_col_idx, value=f"ERROR: invalid subset '{subset}'")
            continue

        base = str(row[base_col]).strip()
        folder = str(row[folder_col]).strip()
        label_raw = row[label_col]
        need = int(row["__need__"] or 0)
        if need <= 0:
            ws.cell(row=header_row + 1 + idx, column=log_col_idx, value="SKIP: 枚数<=0")
            continue

        src_folder = resolve_src_folder(base, folder)
        if not src_folder.exists() or not src_folder.is_dir():
            ws.cell(row=header_row + 1 + idx, column=log_col_idx, value=f"ERROR: src not found: {src_folder}")
            continue

        label_name = sanitize_label(label_raw)
        dst_dir = dest_root / subset / label_name
        dst_dir.mkdir(parents=True, exist_ok=True)
        dest_key = (subset, label_name)

        if not dest_existing_names[dest_key]:
            dest_existing_names[dest_key] = {p.name for p in dst_dir.glob("*.png")}

        pngs = list_pngs(src_folder, recursive=args.recursive)
        candidates: List[SrcFile] = []
        for p in pngs:
            name = p.name
            if name in dest_existing_names[dest_key] or name in dest_reserved_names[dest_key]:
                continue
            store = extract_store(p, store_pat)
            candidates.append(SrcFile(p, store))

        selected = pick_files_balanced_then_fill(candidates, need, args.primary_store_cap, dest_store_counts[dest_key])

        copied = 0
        sample_logs: List[str] = []

        for sf in selected:
            target = dst_dir / sf.path.name
            dest_reserved_names[dest_key].add(sf.path.name)
            if args.dry_run:
                status = "DRY-RUN"
            else:
                try:
                    shutil.copy2(sf.path, target)
                    status = "COPIED"
                    dest_existing_names[dest_key].add(sf.path.name)
                except Exception as e:
                    status = f"ERROR({e})"
            if status.startswith("COPIED") or status == "DRY-RUN":
                copied += 1
                sample_logs.append(f"{sf.path.name}(store={sf.store})[{status}]")

        unmet = max(0, need - copied)
        log_msg = f"{src_folder} -> {dst_dir} need={need} copied={copied} unmet={unmet}"
        if sample_logs:
            log_msg += " | " + " ; ".join(sample_logs[:6])
        ws.cell(row=header_row + 1 + idx, column=log_col_idx, value=log_msg)

    wb.save(excel_path)
    print(f"Done. Wrote logs to column G on sheet '{args.sheet}'.")

if __name__ == "__main__":
    main()

python copy_select_by_excel.py \
  --excel /path/to/book.xlsx \
  --sheet Sheet1 \
  --dest /srv/datasets \
  --primary_store_cap 6 \
  --filename_regex '(?P<store>\d{4})' \
  --dry-run

