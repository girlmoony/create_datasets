#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import argparse
import datetime as dt
import re
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

VALID_SUBSETS = {"train","val","test"}

def sanitize_label(label: str) -> str:
    """
    ラベル文字列をディレクトリ名として安全な形に整形。
    - 先頭末尾の空白を除去
    - OSで問題になりやすい文字を '_' に置換（/ \ : * ? " < > | など）
    - 連続する空白を単一の空白に圧縮
    - 長さ0になった場合は 'unnamed' を返す
    """
    s = str(label or "").strip()
    s = re.sub(r'[\\/:\*\?"<>\|\r\n\t]', '_', s)
    s = re.sub(r'\s+', ' ', s)
    return s if s else "unnamed"

def resolve_src_path(base: str, folder: str) -> Path:
    p = Path(folder)
    if p.is_absolute():
        return p
    return (Path(base) / folder).resolve()

def copy_with_overwrite(src: Path, dst: Path, dry_run: bool) -> str:
    """
    既にdstが存在する場合は削除してからコピー（上書き）。
    """
    if dry_run:
        return "DRY-RUN"
    try:
        if dst.exists():
            shutil.rmtree(dst)
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copytree(src, dst)
        return "COPIED(overwrite)"
    except Exception as e:
        return f"ERROR({e})"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Excelファイルパス（上書き保存されます）")
    ap.add_argument("--sheet", required=True, help="読み書き対象シート名")
    ap.add_argument("--dest", required=True, help="コピー先ベース（直下に train/val/test を作成）")
    ap.add_argument("--dry-run", action="store_true", help="コピーを実行せずログのみ書く")
    args = ap.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")
    dest_root = Path(args.dest).resolve()

    # 読み込み（header=0 を想定）
    df = pd.read_excel(excel_path, sheet_name=args.sheet, header=0)
    cols_lower = {c.lower(): c for c in df.columns}
    required = ["type","product","label","base","subset","folder"]
    missing = [c for c in required if c not in cols_lower]
    if missing:
        raise ValueError(f"必要カラムが見つかりません: {missing}  （存在カラム: {list(df.columns)}）")

    # type==OK のみ
    type_col = cols_lower["type"]
    subset_col = cols_lower["subset"]
    label_col = cols_lower["label"]

    df_ok = df[df[type_col].astype(str).str.upper() == "OK"].copy()
    # subset 正規化
    df_ok[subset_col] = df_ok[subset_col].astype(str).str.strip().str.lower()

    # Excelにログ出力（G列固定）するため openpyxl でロード
    from openpyxl.utils import get_column_letter
    wb = load_workbook(excel_path)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"シートが存在しません: {args.sheet}")
    ws = wb[args.sheet]

    header_row = 1
    log_col_index = 7  # G列

    for idx, row in df_ok.iterrows():
        base = str(row[cols_lower["base"]]).strip()
        subset = str(row[subset_col]).strip()
        folder = str(row[cols_lower["folder"]]).strip()
        label_raw = row[label_col]

        if subset not in VALID_SUBSETS:
            log = f"ERROR: invalid subset '{subset}' (use train/val/test)"
            ws.cell(row=header_row + 1 + idx, column=log_col_index, value=log)
            continue

        src = resolve_src_path(base, folder)
        if not src.exists() or not src.is_dir():
            log = f"ERROR: src not found: {src}"
            ws.cell(row=header_row + 1 + idx, column=log_col_index, value=log)
            continue

        # コピー先：--dest/<subset>/<label>
        label_name = sanitize_label(label_raw)
        dst = dest_root / subset / label_name

        status = copy_with_overwrite(src, dst, args.dry_run)
        log = f"{src} -> {dst} [{status}]"
        ws.cell(row=header_row + 1 + idx, column=log_col_index, value=log)

    wb.save(excel_path)
    print(f"Done. Logs written to column G on sheet '{args.sheet}'.")

if __name__ == "__main__":
    main()python copy_by_excel.py \
  --excel /path/to/book.xlsx \
  --sheet Sheet1 \
  --dest /srv/datasets \
  --dry-run
